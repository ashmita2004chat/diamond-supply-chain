import os
import re
from typing import Dict, Tuple, Optional, List

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
from openpyxl import load_workbook

# =========================================================
# Global page config + CSS (ONLY ONCE)
# =========================================================
st.set_page_config(
    page_title="Diamonds — Trade + Production",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded",
)

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"]  { font-family: 'Inter', sans-serif; }

/* App background (lighter) */
.stApp {
  background: radial-gradient(1200px 600px at 10% 0%, #eef6ff 0%, transparent 60%),
              radial-gradient(1200px 600px at 90% 0%, #fff1e8 0%, transparent 55%),
              linear-gradient(180deg, #fbfcff 0%, #ffffff 60%, #fbfbfd 100%);
}

/* Sidebar background (light) */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #f7faff 0%, #ffffff 45%, #faf7ff 100%);
  border-right: 1px solid rgba(0,0,0,0.06);
}
section[data-testid="stSidebar"] .stMarkdown { font-size: 0.95rem; }

/* Headline card */
.hero {
  padding: 22px 24px;
  border-radius: 18px;
  background: rgba(255,255,255,0.75);
  border: 1px solid rgba(0,0,0,0.06);
  box-shadow: 0 12px 30px rgba(16,24,40,0.08);
  backdrop-filter: blur(8px);
}
.hero h1 { margin: 0; font-size: 2.2rem; }
.hero p { margin: 8px 0 0 0; color: rgba(0,0,0,0.65); }

/* Metric cards spacing */
div[data-testid="stMetric"] {
  background: rgba(255,255,255,0.72);
  border: 1px solid rgba(0,0,0,0.06);
  padding: 14px 14px;
  border-radius: 16px;
  box-shadow: 0 10px 24px rgba(16,24,40,0.06);
}

/* Tabs look */
.stTabs [data-baseweb="tab-list"] button { font-weight: 600; }
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# =========================================================
# Dashboard selector (prevents sidebar mixing)
# =========================================================
st.sidebar.markdown("## Dashboard")
dashboard = st.sidebar.radio(
    "Choose view",
    ["Trade (Diamonds 7102)", "Production (Diamonds)"],
    index=0,
    label_visibility="collapsed",
)

# =========================================================
# ===================== TRADE DASHBOARD ====================
# =========================================================

# Helpers: HS6 → Segment/Subtype mapping
HS6_META = {
    "710210": ("Industrial", "Unworked Unsorted"),
    "710221": ("Industrial", "Unworked Sorted"),
    "710229": ("Industrial", "Worked/Crushed"),
    "710231": ("Non-Industrial", "Unworked"),
    "710239": ("Non-Industrial", "Worked"),
}

def normalize_hs6(x) -> Optional[str]:
    if x is None:
        return None
    s = str(x).strip()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\s+", "", s)
    if re.fullmatch(r"\d{6}", s):
        return s
    return None

def find_row_index(ws, target: str, search_rows: int = 200) -> Optional[int]:
    tgt = target.strip().lower()
    for r in range(1, min(search_rows, ws.max_row) + 1):
        for c in range(1, min(25, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == tgt:
                return r
    return None

def extract_product_line(ws) -> str:
    for r in range(1, 20):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "product" in v.lower():
                txt = " ".join(v.split())
                return txt.replace("Product:", "Product:").strip()
    return ""

def parse_trade_block(ws, hs6: str, block: str) -> pd.DataFrame:
    if block == "imports":
        header_key = "Importers"
        flow = "Imports"
        year_prefix = "Imported value in "
    else:
        header_key = "Exporters"
        flow = "Exports"
        year_prefix = "Exported value in "

    header_row = find_row_index(ws, header_key)
    if header_row is None:
        return pd.DataFrame(columns=["hs6", "flow", "country", "year", "value_usd_thousand"])

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        headers.append(None if v is None else str(v).strip())

    year_cols: List[Tuple[int, int]] = []
    country_col = None
    for idx, h in enumerate(headers, start=1):
        if h == header_key:
            country_col = idx
        if isinstance(h, str) and h.startswith(year_prefix):
            y = h.replace(year_prefix, "").strip()
            if y.isdigit():
                year_cols.append((idx, int(y)))

    if country_col is None or not year_cols:
        return pd.DataFrame(columns=["hs6", "flow", "country", "year", "value_usd_thousand"])

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        country = ws.cell(r, country_col).value
        if country is None or (isinstance(country, str) and country.strip() == ""):
            break
        country = str(country).strip()
        for col_idx, year in year_cols:
            val = ws.cell(r, col_idx).value
            rows.append((hs6, flow, country, year, val))
        r += 1

    df = pd.DataFrame(rows, columns=["hs6", "flow", "country", "year", "value_usd_thousand"])
    df["value_usd_thousand"] = pd.to_numeric(df["value_usd_thousand"], errors="coerce")
    df = df.dropna(subset=["value_usd_thousand"])
    df["country"] = df["country"].astype(str).str.strip()
    return df

@st.cache_data(show_spinner=False)
def load_all_trade_data(xlsx_path: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    descriptions: Dict[str, str] = {}
    all_parts = []

    for sheet_name in wb.sheetnames:
        hs6 = normalize_hs6(sheet_name)
        if not hs6:
            continue
        ws = wb[sheet_name]
        descriptions[hs6] = extract_product_line(ws)

        imp = parse_trade_block(ws, hs6, "imports")
        exp = parse_trade_block(ws, hs6, "exports")

        if not imp.empty:
            all_parts.append(imp)
        if not exp.empty:
            all_parts.append(exp)

    if not all_parts:
        return (
            pd.DataFrame(columns=["hs6", "flow", "country", "year", "value_usd_thousand", "segment", "subtype"]),
            descriptions
        )

    df = pd.concat(all_parts, ignore_index=True)

    # meticulous de-duplication
    df = df.groupby(["hs6", "flow", "year", "country"], as_index=False)["value_usd_thousand"].sum()

    df["segment"] = df["hs6"].map(lambda x: HS6_META.get(x, ("Unknown", "Unknown"))[0])
    df["subtype"] = df["hs6"].map(lambda x: HS6_META.get(x, ("Unknown", "Unknown"))[1])

    return df, descriptions

def format_value(v: float, unit: str) -> str:
    return f"{v:,.2f}" if unit == "USD Mn" else f"{v:,.0f}"

def render_trade_tab():
    st.markdown(
        """
    <div class="hero">
      <h1>💎 Diamonds (HS 7102) — Trade Supply Chain</h1>
      <p>Interactive filters • Top partners • Trend view • Imports/Exports split • Cleaned + de-duplicated</p>
    </div>
    """,
        unsafe_allow_html=True,
    )
    st.write("")

    st.sidebar.markdown("## Trade Filters")

    DEFAULT_FILE = "Diamonds(7102).xlsx"

    with st.sidebar.expander("Trade data source", expanded=True):
        src = st.radio(
            "Choose data source",
            ["Auto-load default file", "Upload a file"],
            index=0,
            help=f"Auto-load looks for `{DEFAULT_FILE}` in the same folder as app.py."
        )

    uploaded = None
    xlsx_path = None

    if src == "Upload a file":
        uploaded = st.sidebar.file_uploader("Upload Diamonds(7102).xlsx", type=["xlsx"], key="trade_uploader")
        if uploaded is not None:
            xlsx_path = uploaded
    else:
        if os.path.exists(DEFAULT_FILE):
            xlsx_path = DEFAULT_FILE
        else:
            st.sidebar.warning(f"Default file not found: `{DEFAULT_FILE}`. Upload instead.")
            uploaded = st.sidebar.file_uploader("Upload Diamonds(7102).xlsx", type=["xlsx"], key="trade_uploader_fallback")
            if uploaded is not None:
                xlsx_path = uploaded

    unit = st.sidebar.radio("Value unit", ["USD Mn", "USD Thousand"], index=0, key="trade_unit")

    if xlsx_path is None:
        st.info("Upload the Trade Excel file (or keep it as `Diamonds(7102).xlsx` next to app.py) to start.")
        st.stop()

    with st.spinner("Loading + cleaning trade workbook..."):
        df, hs6_desc = load_all_trade_data(xlsx_path)

    if df.empty:
        st.error("No HS6 sheets parsed in Trade file. Please verify the Excel structure.")
        st.stop()

    df = df.copy()
    df["value"] = df["value_usd_thousand"] / 1000.0 if unit == "USD Mn" else df["value_usd_thousand"]

    segments = sorted(df["segment"].dropna().unique().tolist())
    flows = ["Imports", "Exports"]
    hs6_list = sorted(df["hs6"].unique().tolist(), key=lambda x: int(x))
    years = sorted(df["year"].unique().tolist())

    with st.sidebar.expander("Trade filters", expanded=True):
        seg = st.selectbox("Segment", segments, index=0 if segments else 0, key="trade_seg")
        subtypes = sorted(df.loc[df["segment"] == seg, "subtype"].unique().tolist())
        subtype = st.selectbox("Subtype", subtypes, index=0 if subtypes else 0, key="trade_subtype")

        hs6_filtered = sorted(
            df.loc[(df["segment"] == seg) & (df["subtype"] == subtype), "hs6"].unique().tolist(),
            key=lambda x: int(x)
        )
        hs6 = st.selectbox("HS6", hs6_filtered if hs6_filtered else hs6_list, index=0, key="trade_hs6")

        flow = st.selectbox("Flow (Overview)", flows, index=0, key="trade_flow_overview")
        year = st.slider("Year (Overview)", min_value=min(years), max_value=max(years), value=min(years), step=1, key="trade_year")
        top_n = st.slider("Top N partners", min_value=3, max_value=25, value=8, step=1, key="trade_topn")

        include_world = st.checkbox("Include 'World' row", value=False, key="trade_world")

    base = df[
        (df["segment"] == seg) &
        (df["subtype"] == subtype) &
        (df["hs6"] == hs6) &
        (df["flow"] == flow) &
        (df["year"] == year)
    ].copy()

    if not include_world:
        base = base[base["country"].str.lower() != "world"]

    base = base.groupby(["country"], as_index=False)["value"].sum()

    tab1, tab2, tab3 = st.tabs(["📌 Overview", "📈 Trend", "🧾 Data"])

    with tab1:
        desc_line = hs6_desc.get(hs6, "").strip()
        if desc_line:
            st.caption(desc_line)

        total_val = float(base["value"].sum()) if not base.empty else 0.0
        partner_count = int(base["country"].nunique()) if not base.empty else 0

        if not base.empty:
            top_row = base.sort_values("value", ascending=False).iloc[0]
            top_country = str(top_row["country"])
            top_value = float(top_row["value"])
        else:
            top_country, top_value = "-", 0.0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric(f"Total ({unit})", format_value(total_val, unit))
        c2.metric("Top Country", top_country)
        c3.metric(f"Top Value ({unit})", format_value(top_value, unit))
        c4.metric("Partner count", f"{partner_count:,}")

        st.write("")

        top_df = base.sort_values("value", ascending=False).head(top_n).copy()
        top_df["rank"] = np.arange(1, len(top_df) + 1)
        chart_df = top_df.sort_values("value", ascending=True)

        left, right = st.columns([1.2, 1.0], gap="large")

        with left:
            st.subheader(f"Top {len(top_df)} Partners — {flow} ({year})")
            fig = px.bar(
                chart_df,
                x="value",
                y="country",
                orientation="h",
                text="value",
                labels={"value": f"Value ({unit})", "country": "Country"},
            )
            fig.update_traces(texttemplate="%{text:.2f}" if unit == "USD Mn" else "%{text:.0f}", textposition="inside")
            fig.update_layout(
                template="plotly_white",
                height=420,
                margin=dict(l=10, r=10, t=10, b=10),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(fig, use_container_width=True)

        with right:
            st.subheader("Snapshot Table (Top N)")
            table_df = top_df[["rank", "country", "value"]].rename(columns={"value": f"Value ({unit})"})
            st.dataframe(table_df, use_container_width=True, height=420)

    with tab2:
        st.subheader("Trend (All Years) — Selected HS6")

        trend_countries = df[(df["hs6"] == hs6) & (df["segment"] == seg) & (df["subtype"] == subtype)].copy()
        if not include_world:
            trend_countries = trend_countries[trend_countries["country"].str.lower() != "world"]

        country_for_trend = st.selectbox(
            "Country for trend",
            sorted(trend_countries["country"].unique().tolist()),
            index=0,
            key="trade_country_trend"
        )

        flow_for_trend = st.selectbox(
            "Flow for trend (separate from Overview)",
            ["Imports", "Exports"],
            index=0,
            key="trade_flow_trend"
        )

        tr = df[
            (df["hs6"] == hs6) &
            (df["segment"] == seg) &
            (df["subtype"] == subtype) &
            (df["country"] == country_for_trend) &
            (df["flow"] == flow_for_trend)
        ].copy()

        tr = tr.groupby(["year"], as_index=False)["value"].sum().sort_values("year")

        if tr.empty:
            st.info("No trend data for this selection.")
        else:
            fig2 = px.line(
                tr,
                x="year",
                y="value",
                markers=True,
                labels={"value": f"Value ({unit})", "year": "Year"},
            )
            fig2.update_layout(template="plotly_white", height=460, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig2, use_container_width=True)
            st.caption("Tip: Switch Imports/Exports above without changing your Overview filters.")

    with tab3:
        st.subheader("Filtered dataset (downloadable)")
        data_view = df[(df["segment"] == seg) & (df["subtype"] == subtype) & (df["hs6"] == hs6)].copy()
        if not include_world:
            data_view = data_view[data_view["country"].str.lower() != "world"]

        keep_cols = ["segment", "subtype", "hs6", "flow", "country", "year", "value"]
        data_view = data_view[keep_cols].sort_values(["flow", "year", "country"])

        st.dataframe(data_view, use_container_width=True, height=520)
        csv = data_view.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download CSV (filtered)",
            data=csv,
            file_name=f"diamonds_7102_{hs6}_{seg}_{subtype}.csv".replace(" ", "_").lower(),
            mime="text/csv",
        )

    st.caption("Trade note: Values are parsed from Diamonds(7102).xlsx and de-duplicated at (HS6, Flow, Year, Country).")

# =========================================================
# ================== PRODUCTION DASHBOARD ==================
# =========================================================

DEFAULT_PROD_FILE = "Production of Diamonds.xlsx"

def hs_to_hs6(x) -> Optional[str]:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    # Handle numeric like 7102.31 (float) and keep 2 decimals
    if isinstance(x, (int, float, np.integer, np.floating)):
        s = f"{float(x):.2f}"
    else:
        s = str(x).strip().replace(" ", "")
    if re.fullmatch(r"\d+\.\d+", s):
        left, right = s.split(".")
        right = (right + "00")[:2]
        hs6 = left + right
        return hs6 if re.fullmatch(r"\d{6}", hs6) else None
    if re.fullmatch(r"\d{6}", s):
        return s
    return None

@st.cache_data(show_spinner=False)
def load_production_df_from_path(path: str) -> pd.DataFrame:
    usecols = ["Country Name", "Year", "Quarter", "Trade Type", "Origin", "Prov_Dest", "HS Code", "Carat", "US Value"]
    dfp = pd.read_excel(path, sheet_name=0, usecols=usecols, engine="openpyxl")
    return _prep_production(dfp)

@st.cache_data(show_spinner=False)
def load_production_df_from_upload(uploaded) -> pd.DataFrame:
    usecols = ["Country Name", "Year", "Quarter", "Trade Type", "Origin", "Prov_Dest", "HS Code", "Carat", "US Value"]
    dfp = pd.read_excel(uploaded, sheet_name=0, usecols=usecols, engine="openpyxl")
    return _prep_production(dfp)

def _prep_production(dfp: pd.DataFrame) -> pd.DataFrame:
    dfp = dfp.rename(columns={
        "Country Name": "Reporter",
        "Trade Type": "Flow",
        "Prov_Dest": "Partner",
        "HS Code": "HS_raw",
        "US Value": "US_Value",
    }).copy()

    dfp["HS6"] = dfp["HS_raw"].apply(hs_to_hs6)
    dfp = dfp.dropna(subset=["Reporter", "Year", "Quarter", "Flow", "Origin", "Partner", "HS6"])

    dfp["Year"] = pd.to_numeric(dfp["Year"], errors="coerce").astype("Int64")
    dfp["Quarter"] = pd.to_numeric(dfp["Quarter"], errors="coerce").astype("Int64")
    dfp["Carat"] = pd.to_numeric(dfp["Carat"], errors="coerce")
    dfp["US_Value"] = pd.to_numeric(dfp["US_Value"], errors="coerce")

    dfp = dfp.dropna(subset=["Year", "Quarter"])
    dfp = dfp[(dfp["Year"] >= 1900) & (dfp["Year"] <= 2100)]
    dfp["Flow"] = dfp["Flow"].astype(str).str.strip().str.title()
    dfp["Flow"] = dfp["Flow"].replace({"Export": "Exports", "Import": "Imports"})

    dfp["US_Value_Mn"] = dfp["US_Value"] / 1_000_000.0
    dfp["USD_per_Carat"] = np.where(dfp["Carat"] > 0, dfp["US_Value"] / dfp["Carat"], np.nan)

    return dfp

def _metric_col(metric: str) -> str:
    return {"Carat": "Carat", "US Value (USD Mn)": "US_Value_Mn", "USD per Carat": "USD_per_Carat"}[metric]

def render_production_tab():
    st.markdown(
        """
    <div class="hero">
      <h1>💎 Diamonds — Production Analytics</h1>
      <p>This tab uses ONLY Production of Diamonds.xlsx (independent from the Trade dashboard)</p>
    </div>
    """,
        unsafe_allow_html=True,
    )
    st.write("")

    st.sidebar.markdown("## Production Filters")

    with st.sidebar.expander("Production data source", expanded=True):
        prod_src = st.radio(
            "Choose data source",
            ["Auto-load default file", "Upload a file"],
            index=0,
            key="prod_src"
        )

    prod_uploaded = None
    if prod_src == "Upload a file":
        prod_uploaded = st.sidebar.file_uploader("Upload Production of Diamonds.xlsx", type=["xlsx"], key="prod_uploader")

    if prod_src == "Auto-load default file":
        if not os.path.exists(DEFAULT_PROD_FILE):
            st.warning(f"Default production file not found: `{DEFAULT_PROD_FILE}`. Upload it from the sidebar.")
            return
        with st.spinner("Loading production data..."):
            dfp = load_production_df_from_path(DEFAULT_PROD_FILE)
    else:
        if prod_uploaded is None:
            st.info("Upload the Production Excel file to start.")
            return
        with st.spinner("Loading production data..."):
            dfp = load_production_df_from_upload(prod_uploaded)

    # Sidebar filters (production only)
    reporters = sorted(dfp["Reporter"].unique().tolist())
    reporter = st.sidebar.selectbox("Reporter", reporters, index=0, key="prod_reporter")

    hs6_list = sorted(dfp["HS6"].unique().tolist())
    hs6 = st.sidebar.selectbox("HS6 (from HS Code)", hs6_list, index=0, key="prod_hs6")

    flows = ["All"] + sorted(dfp["Flow"].unique().tolist())
    flow = st.sidebar.selectbox("Flow", flows, index=0, key="prod_flow")

    y_min, y_max = int(dfp["Year"].min()), int(dfp["Year"].max())
    yr0, yr1 = st.sidebar.slider("Year range", y_min, y_max, (y_min, y_max), key="prod_years")

    q_opts = ["All", "1", "2", "3", "4"]
    quarter = st.sidebar.selectbox("Quarter", q_opts, index=0, key="prod_quarter")

    metric = st.sidebar.radio("Metric", ["Carat", "US Value (USD Mn)", "USD per Carat"], index=1, key="prod_metric")
    top_n = st.sidebar.slider("Top N origins", 5, 25, 10, key="prod_topn")

    # Apply filters
    d = dfp[
        (dfp["Reporter"] == reporter) &
        (dfp["HS6"] == hs6) &
        (dfp["Year"].between(yr0, yr1, inclusive="both"))
    ].copy()

    if flow != "All":
        d = d[d["Flow"] == flow]

    if quarter != "All":
        d = d[d["Quarter"] == int(quarter)]

    # Tabs inside production
    t1, t2, t3, t4 = st.tabs(["📌 Overview", "📈 Trends", "🔁 Flows", "🧾 Data"])

    # ---------- Overview ----------
    with t1:
        if d.empty:
            st.info("No production rows for the selected filters.")
        else:
            total_carat = float(d["Carat"].sum(skipna=True))
            total_val_mn = float(d["US_Value_Mn"].sum(skipna=True))
            avg_usd_carat = float(np.nanmean(d["USD_per_Carat"]))
            n_origins = int(d["Origin"].nunique())
            n_partners = int(d["Partner"].nunique())

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Total Carats", f"{total_carat:,.0f}")
            c2.metric("Total Value (USD Mn)", f"{total_val_mn:,.2f}")
            c3.metric("Avg USD/Carat", f"{avg_usd_carat:,.2f}")
            c4.metric("#Origins", f"{n_origins:,}")
            c5.metric("#Destinations", f"{n_partners:,}")

            st.write("")

            mcol = _metric_col(metric)

            # METICULOUS aggregation (prevents any duplication in visuals)
            if metric == "USD per Carat":
                agg = d.groupby("Origin", as_index=False).agg(Carat=("Carat", "sum"), US_Value=("US_Value", "sum"), US_Value_Mn=("US_Value_Mn", "sum"))
                agg["USD_per_Carat"] = np.where(agg["Carat"] > 0, agg["US_Value"] / agg["Carat"], np.nan)
            else:
                agg = d.groupby("Origin", as_index=False).agg(
                    Carat=("Carat", "sum"),
                    US_Value_Mn=("US_Value_Mn", "sum"),
                    USD_per_Carat=("USD_per_Carat", "mean"),
                )

            agg = agg.sort_values(mcol, ascending=False).head(top_n).reset_index(drop=True)
            agg["Rank"] = np.arange(1, len(agg) + 1)

            left, right = st.columns([1.2, 1.0], gap="large")

            with left:
                chart_df = agg.sort_values(mcol, ascending=True)
                fig = px.bar(
                    chart_df,
                    x=mcol,
                    y="Origin",
                    orientation="h",
                    text=mcol,
                    labels={mcol: metric, "Origin": "Origin"},
                    title=f"Top {len(agg)} Origins — {metric} (Largest → Smallest)",
                )
                fig.update_layout(template="plotly_white", height=520, margin=dict(l=10, r=10, t=60, b=10), yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig, use_container_width=True)

            with right:
                show = agg[["Rank", "Origin", mcol]].copy()
                show = show.rename(columns={mcol: metric})
                st.subheader("Snapshot Table (Top N)")
                st.dataframe(show, use_container_width=True, height=520)

            # Concentration insight (HHI) by year
            with st.expander("Show concentration (HHI) trend"):
                by = d.groupby(["Year", "Origin"], as_index=False)["US_Value"].sum()
                total_y = by.groupby("Year", as_index=False)["US_Value"].sum().rename(columns={"US_Value": "Total"})
                by = by.merge(total_y, on="Year", how="left")
                by["share"] = np.where(by["Total"] > 0, by["US_Value"] / by["Total"], 0)
                hhi = by.groupby("Year", as_index=False).apply(lambda x: float((x["share"] ** 2).sum())).rename(columns={None: "HHI"})
                hhi["HHI"] = hhi["HHI"] * 10000  # standard scaling
                fig_hhi = px.line(hhi.sort_values("Year"), x="Year", y="HHI", markers=True, title="HHI (Origin concentration) over time")
                fig_hhi.update_layout(template="plotly_white", height=420, margin=dict(l=10, r=10, t=60, b=10))
                st.plotly_chart(fig_hhi, use_container_width=True)

    # ---------- Trends ----------
    with t2:
        if d.empty:
            st.info("No production rows for the selected filters.")
        else:
            origins = sorted(d["Origin"].unique().tolist())
            origin_sel = st.selectbox("Origin for trend", origins, index=0, key="prod_origin_trend")
            flow_sel = st.selectbox("Flow for trend", ["All", "Imports", "Exports"], index=0, key="prod_flow_trend")
            metric_t = st.radio("Trend metric", ["Carat", "US Value (USD Mn)", "USD per Carat"], index=1, horizontal=True, key="prod_metric_trend")
            mcol = _metric_col(metric_t)

            dt = d[d["Origin"] == origin_sel].copy()
            if flow_sel != "All":
                dt = dt[dt["Flow"] == flow_sel]

            yearly = dt.groupby("Year", as_index=False).agg(
                Carat=("Carat", "sum"),
                US_Value=("US_Value", "sum"),
                US_Value_Mn=("US_Value_Mn", "sum")
            )
            yearly["USD_per_Carat"] = np.where(yearly["Carat"] > 0, yearly["US_Value"] / yearly["Carat"], np.nan)

            fig = px.line(yearly.sort_values("Year"), x="Year", y=mcol, markers=True, title=f"{origin_sel} — {metric_t} trend")
            fig.update_layout(template="plotly_white", height=460, margin=dict(l=10, r=10, t=60, b=10))
            st.plotly_chart(fig, use_container_width=True)

            with st.expander("Show seasonality heatmap (Year × Quarter)"):
                q = dt.groupby(["Year", "Quarter"], as_index=False).agg(
                    Carat=("Carat", "sum"),
                    US_Value=("US_Value", "sum"),
                    US_Value_Mn=("US_Value_Mn", "sum")
                )
                q["USD_per_Carat"] = np.where(q["Carat"] > 0, q["US_Value"] / q["Carat"], np.nan)
                piv = q.pivot(index="Year", columns="Quarter", values=mcol).fillna(0)
                fig_hm = px.imshow(piv, aspect="auto", title=f"Seasonality heatmap — {metric_t} ({origin_sel})")
                fig_hm.update_layout(template="plotly_white", height=520, margin=dict(l=10, r=10, t=60, b=10))
                st.plotly_chart(fig_hm, use_container_width=True)

    # ---------- Flows ----------
    with t3:
        if d.empty:
            st.info("No production rows for the selected filters.")
        else:
            origins = sorted(d["Origin"].unique().tolist())
            origin_sel = st.selectbox("Choose Origin", origins, index=0, key="prod_origin_flow")

            metric_f = st.radio("Flow metric", ["US Value (USD Mn)", "Carat"], index=0, horizontal=True, key="prod_metric_flow")
            mcol = "US_Value_Mn" if metric_f == "US Value (USD Mn)" else "Carat"

            dd = d[d["Origin"] == origin_sel].copy()
            top_dest = dd.groupby("Partner", as_index=False)[mcol].sum().sort_values(mcol, ascending=False).head(20)

            fig = px.bar(
                top_dest.sort_values(mcol, ascending=True),
                x=mcol,
                y="Partner",
                orientation="h",
                title=f"Top Destinations for {origin_sel} — {metric_f}",
            )
            fig.update_layout(template="plotly_white", height=520, margin=dict(l=10, r=10, t=60, b=10))
            st.plotly_chart(fig, use_container_width=True)

    # ---------- Data ----------
    with t4:
        if d.empty:
            st.info("No production rows for the selected filters.")
        else:
            agg = d.groupby(["Origin", "Partner", "Year", "Quarter", "Flow"], as_index=False).agg(
                Carat=("Carat", "sum"),
                US_Value=("US_Value", "sum"),
                US_Value_Mn=("US_Value_Mn", "sum"),
            )
            agg["USD_per_Carat"] = np.where(agg["Carat"] > 0, agg["US_Value"] / agg["Carat"], np.nan)
            agg = agg.sort_values(["Year", "US_Value_Mn"], ascending=[True, False])

            st.dataframe(agg, use_container_width=True, height=520)

            csv = agg.to_csv(index=False).encode("utf-8")
            st.download_button("Download aggregated CSV", csv, file_name="production_diamonds_aggregated.csv", mime="text/csv")

    st.caption("Production note: Built only from Production of Diamonds.xlsx, aggregated carefully to prevent duplicates.")

# =========================================================
# Run selected dashboard
# =========================================================
if dashboard.startswith("Trade"):
    render_trade_tab()
else:
    render_production_tab()
